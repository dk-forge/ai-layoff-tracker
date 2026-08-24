"""
Custom WARN collectors for states the open-source warn-scraper can no longer
read (their sites changed). Each fetcher was built from a machine-verified
recon spec (2026-07-15) and returns entries in the exact same shape — and with
the exact same dedup-hash formula — as sources/warn.py, so re-imports upsert
instead of duplicating.

States: TX (Socrata JSON), FL (POST form -> HTML table), GA (WP admin-ajax),
OH (per-year CSVs on the Ohio DAM, archives to 2020), MI (Sitecore search
JSON), CO (Google Sheets xlsx), ID (text PDF), LA (per-year text PDFs, 2015+
via Wayback for the years laworks.net dropped), NC (per-year CSV + archive
PDFs to 2015), NV (master PDFs), MN (report PDFs), MA (CSV + FY xlsx), and
NY history (the retired dol.ny.gov database, frozen 4/1/2025 — supplements
warn-scraper's live-dashboard NY coverage; identical hashes dedup the overlap).
"""
import hashlib
import html as _html
import io
import json
import re

import requests

from .warn import _count, _to_iso_date, STATE_WARN_URL
from .warn_llm import llm_count_from_text

UA = {"User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                     "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36")}
TIMEOUT = 40


def _entry(st, company, jobs, date, city="", kind="", detail_url=""):
    """Normalized entry — MUST mirror sources/warn.py exactly (incl. the hash)."""
    company = (company or "").strip()
    city = (city or "").strip()
    if not company or jobs <= 0 or jobs > 100000 or not date:
        return None
    if date < "2015-01-01" or date > "2028-12-31":
        return None
    loc = f" in {city}" if city else ""
    excerpt = (f"{kind or 'Layoff'} at {company}{loc}. "
               f"{jobs:,} employees affected, effective {date}. "
               f"Filed under the {st} WARN Act.")
    # City is deliberately NOT hashed: collectors emit inconsistent city text for
    # the same notice (blank vs "Seattle and Remote" vs a county), which spawned
    # cross-collector orphan duplicates. company+date+jobs+state identifies it.
    hash_input = f"warn{company.lower().strip()}{date}{jobs}{st}"
    return {
        "source_type": "warn",
        "source_name": f"{st} WARN notice",
        "verification_level": "warn",
        "company_name": company,
        "ticker": None,
        "job_count": jobs,
        "layoff_date": date,
        "industry": None,
        "country": "United States",
        "state": st,
        "roles": None,
        "excerpt": excerpt,
        # Tag permanent shutdowns so "Closures only" is filterable. dedup_hash
        # excludes reason_tags, so this never affects dedup/upsert.
        "reason_tags": (["closure"] if kind and "clos" in kind.lower() else []),
        "ai_explicit": False,
        "ai_language": None,
        "source_url": detail_url or STATE_WARN_URL.get(st, ""),
        "dedup_hash": hashlib.md5(hash_input.encode("utf-8")).hexdigest(),
        "is_layoff_event": True,
    }


def _strip_tags(markup):
    return _html.unescape(re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", markup or "")).strip())


def fetch_tx():
    """Official TWC Socrata dataset (full history 2019+, updated ~weekly)."""
    url = ("https://data.texas.gov/resource/8w53-c4f6.json"
           "?$limit=5000&$order=notice_date%20DESC")
    rows = requests.get(url, headers=UA, timeout=TIMEOUT).json()
    out = []
    for r in rows:
        jobs = _count(str(r.get("total_layoff_number") or ""))
        date = _to_iso_date((r.get("layoff_date") or r.get("notice_date") or "")[:10])
        e = _entry("TX", r.get("job_site_name"), jobs, date, r.get("city_name") or "",
                   detail_url="https://data.texas.gov/d/8w53-c4f6")
        if e:
            out.append(e)
    return out


def fetch_fl():
    """FloridaCommerce REACT export: POST returns a fake-Excel HTML table."""
    resp = requests.post(
        "https://reactwarn.floridajobs.org/warnlist/reports",
        data={"StateNotificationStartDate": "2019-01-01",
              "StateNotificationEndDate": "2028-12-31", "appForm": "export"},
        headers=UA, timeout=120)
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", resp.text, re.S | re.I)
    # WarnNo, LWDB, Company, Addr1, Addr2, City, State, Zip, County,
    # NotifDate, LayoffBegin, LayoffEnd, Affected, Industry
    # Florida's export bundles staff TEST notices. Row-level "test" matching was
    # NOT enough: notice W-1511 spans "AT&T" + "BOEING" + "BOEING test" with
    # fabricated counts — one row carried 78,788 fake workers under an AT&T name
    # and briefly topped the whole tracker, while its AT&T/BOEING siblings (clean
    # company cells) slipped through. Quarantine the ENTIRE WarnNo when ANY of its
    # rows is a test row, so the poisoned notice is dropped whole. Legit
    # multi-site notices have no "test" sibling and are untouched.
    test_re = re.compile(r"(?:^|[^a-z])test(?:[^a-z]|\d|$)", re.I)
    parsed = []
    poisoned = set()
    for tr in rows[1:]:  # first <tr> is the header (td cells, no th)
        cells = [_strip_tags(c) for c in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)]
        if len(cells) < 14:
            continue
        parsed.append(cells)
        if test_re.search(cells[2]) and cells[0].strip():
            poisoned.add(cells[0].strip())  # WarnNo of a test-tainted notice
    out = []
    skipped_test = 0
    for cells in parsed:
        warn_no = cells[0].strip()
        if test_re.search(cells[2]) or (warn_no and warn_no in poisoned):
            skipped_test += 1
            continue
        jobs = _count(cells[12])
        date = _to_iso_date(cells[10]) or _to_iso_date(cells[9])
        if jobs <= 0:  # count column may have shifted; let the LLM read the row
            jobs = llm_count_from_text(" ".join(str(c) for c in cells), f"FL {cells[2]}")
        e = _entry("FL", cells[2], jobs, date, cells[5])
        if e:
            out.append(e)
    if skipped_test:
        print(f"    FL: skipped {skipped_test} state-side test row(s) "
              f"across {len(poisoned)} poisoned notice(s)")
    return out


def fetch_ga():
    """TCSG GravityView DataTables endpoint (nonce scraped from the page)."""
    landing = "https://www.tcsg.edu/warn-public-view/"
    page = requests.get(landing, headers=UA, timeout=TIMEOUT).text
    m = re.search(r'"action":"gv_datatables_data"[^}]*?"nonce":"([0-9a-f]{10})"', page)
    if not m:
        m = re.search(r'"nonce":"([0-9a-f]{10})"[^}]*?"action":"gv_datatables_data"', page)
    if not m:
        raise RuntimeError("GA: nonce not found on landing page")
    resp = requests.post(
        "https://www.tcsg.edu/wp-admin/admin-ajax.php",
        data={"action": "gv_datatables_data", "view_id": "77460", "post_id": "77462",
              "nonce": m.group(1), "getData": "false", "draw": "1",
              "start": "0", "length": "-1"},
        headers={**UA, "Referer": landing}, timeout=120)
    data = resp.json().get("data", [])
    out = []
    for row in data:
        # Rows are heterogeneous: plain arrays OR dicts keyed "0".."4"
        if isinstance(row, dict):
            cells = [row.get(str(i)) or row.get(i) or "" for i in range(5)]
        else:
            cells = list(row)
        if len(cells) < 4:
            continue
        link = re.search(r'href="([^"]+)"', str(cells[0]) or "")
        jobs = _count(_strip_tags(str(cells[3])))
        date = _to_iso_date(_strip_tags(str(cells[2])))
        if jobs <= 0:
            jobs = llm_count_from_text(" ".join(_strip_tags(str(c)) for c in cells),
                                       f"GA {_strip_tags(str(cells[1]))}")
        e = _entry("GA", _strip_tags(str(cells[1])), jobs, date,
                   detail_url=link.group(1) if link else "")
        if e:
            out.append(e)
    return out


# Ohio's 2026 site rebuild moved WARN under /job-workforce-services and added
# per-year archive pages. Archives exist for 2020+ only — the 2015-2019 slugs
# 404 (never migrated; verified 2026-07-18). Each page links its year's CSV
# on the Ohio DAM, now with f_auto/q_auto/v<nnn> segments in the path.
_OH_WARN_BASE = ("https://jfs.ohio.gov/job-workforce-services/"
                 "job-programs-and-services/submit-a-warn-notice")
_OH_ARCHIVE_START = 2020
_OH_CSV_RE = re.compile(r'https://dam\.assets\.ohio\.gov/raw/upload/[^"\\ ]+\.csv')


def _oh_entries_from_csv(text):
    """Parse one Ohio DAM CSV (headers may sit a couple of metadata rows down)."""
    import csv as _csv
    lines = text.splitlines()
    start = next((i for i, ln in enumerate(lines[:6]) if "company" in ln.lower()), None)
    if start is None:
        return []
    out = []
    for row in _csv.DictReader(io.StringIO("\n".join(lines[start:]))):
        rl = {(k or "").lower().strip(): (v or "").strip() for k, v in row.items() if k}
        company = rl.get("company") or ""
        jobs = _count(rl.get("potential number affected") or rl.get("number affected") or "")
        date = (_to_iso_date(rl.get("layoff date(s)") or "")
                or _to_iso_date(rl.get("date received") or ""))
        city = (rl.get("city/county") or "").split("/")[0]
        e = _entry("OH", company, jobs, date, city, detail_url=rl.get("url") or "")
        if e:
            out.append(e)
    return out


def fetch_oh():
    """Ohio DAM CSVs discovered from the JFS current page + per-year archive
    pages (2020+). Every page and every CSV is fail-isolated."""
    from datetime import date as _date
    this_year = _date.today().year
    pages = [f"{_OH_WARN_BASE}/current-public-notices-of-layoffs-and-closures"]
    pages += [f"{_OH_WARN_BASE}/{y}-public-notices-of-layoffs-and-closures"
              for y in range(_OH_ARCHIVE_START, this_year)]
    urls = []
    for page_url in pages:
        slug = page_url.rsplit("/", 1)[-1]
        try:
            page = requests.get(page_url, headers=UA, timeout=TIMEOUT)
            if page.status_code != 200:
                print(f"    OH: {slug} -> HTTP {page.status_code}")
                continue
            found = _OH_CSV_RE.findall(page.text)
            if not found:
                print(f"    OH: no DAM CSV link on {slug}")
            urls += found
        except Exception as exc:
            print(f"    OH: {slug} failed ({exc})")
    # Versionless fallback for the years the PAGES did not yield a CSV. The old
    # form guessed {y}/{y}-warn-notice.csv and so only ever resolved for the
    # current year: JFS re-uploads every archive year into the CURRENT year's
    # DAM folder, and the separator flips between years (2024_warn_notice.csv,
    # 2022-warn-notice.csv). With the pages unreachable that guess recovered 61
    # of 787 notices while the scraper still reported a healthy non-zero count.
    # Verified 2026-08-13: this matrix recovers every year but 2023, whose file
    # is an un-guessable one-off (2023-warn-notice_1_9.csv) reachable only via
    # its page. Fallbacks are added ONLY for missing years, so a healthy run
    # (pages up) issues no extra requests beyond the ones that dedup away.
    got_years = {m.group(1) for m in
                 (re.search(r"/(20\d\d)[-_]warn[-_]notice", u, re.I) for u in urls) if m}
    for y in range(_OH_ARCHIVE_START, this_year + 1):
        if str(y) in got_years:
            continue
        for sep in ("-", "_"):
            urls.append("https://dam.assets.ohio.gov/raw/upload/jfs.ohio.gov/"
                        f"{this_year}/{y}{sep}warn{sep}notice.csv")
    out, seen = [], set()
    for u in urls:
        # same file relinked under new renderer flags / version segments
        key = re.sub(r"(?:f_auto/|q_auto/|v\d+/)", "", u)
        if key in seen:
            continue
        seen.add(key)
        try:
            resp = requests.get(u, headers=UA, timeout=TIMEOUT)
            if resp.status_code != 200:
                continue
            out += _oh_entries_from_csv(resp.text)
        except Exception:
            continue
    return out


def fetch_mi():
    """Michigan.gov Sitecore SXA search API: JSON of HTML fragments in two
    shapes — older records use <a class="content-title-link"> + <p> pairs with
    'Commencing date'; newer ones use <h3> + <li> items with 'Layoff date'."""
    url = ("https://www.michigan.gov/leo/sxa/search/results/"
           "?s={8E97AB1D-D2D4-47F8-8CC4-3F1039C8854F}"
           "&itemid={BE81F7C2-36A8-4FDE-853C-B05B6E090055}"
           "&v={1FFFCC21-5151-4A2B-ABFC-F7FE4E5C9783}&p=500"
           "&o=Created%20Date%20sort,Descending")
    data = requests.get(url, headers=UA, timeout=TIMEOUT).json()
    out = []
    for res in data.get("Results", []):
        html = (res.get("Html") or "").replace("&nbsp;", " ")
        # Take the first pattern that yields a NON-EMPTY name, not the first
        # that matches. Newer fragments render the anchor as a placeholder
        # (href="" with no text) and put the real name in the <h3> after it;
        # `or` short-circuits on the match OBJECT, so the empty anchor won and
        # 23 of 112 live records were dropped for an empty company (measured
        # 2026-08-19). No name in either place stays a SKIP - the URL slug is
        # not a company name and must never be used as one.
        title = ""
        for pat in (r'class="content-title-link"[^>]*>(.*?)</a>',
                    r"<h3[^>]*>(.*?)</h3>"):
            m = re.search(pat, html, re.S)
            if m:
                title = _strip_tags(m.group(1))
                if title:
                    break

        def field(label):
            fm = re.search(rf"<strong>\s*{label}[^<]*</strong>\s*:?\s*([^<]+)", html, re.I)
            return fm.group(1).strip() if fm else ""

        jobs = _count(field("Number of jobs impacted"))
        date = (_to_iso_date(field("Layoff date")) or _to_iso_date(field("Commencing date"))
                or _to_iso_date(field("Date of layoff")) or _to_iso_date(field("Date WARN received")))
        if not date:  # labels shift; scan the fragment prose for the first real date
            for cand in re.findall(
                    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|[A-Z][a-z]+\.?\s+\d{1,2},?\s+\d{4}|\d{4}-\d{2}-\d{2})",
                    _strip_tags(html)):
                d = _to_iso_date(cand)
                if d:
                    date = d
                    break
        if jobs <= 0:  # label may have changed; read the whole record fragment
            jobs = llm_count_from_text(_strip_tags(html), f"MI {title}")
        county = field("County")
        link = res.get("Url") or ""
        if link and link.startswith("/"):
            link = "https://www.michigan.gov" + link
        e = _entry("MI", title, jobs, date, county, detail_url=link)
        if e:
            out.append(e)
    return out


def fetch_co(get=None):
    """Colorado CDLE: per-year Google Sheets workbooks linked from the landing page.

    Two things can go dark here, and only ONE of them is our parser's fault, so
    they must not be reported the same way. The CDLE landing page is now behind a
    bot wall (bare HTTP 403), so live workbook-ID discovery is dead and we lean on
    the pinned fallback IDs; that alone is survivable while 2026's workbook stays
    current. But the Google export endpoint also rate-limits a CI runner's IP with
    a 403/429 HTML body — and when every workbook is refused that way, this used
    to return 0 rows silently, which the per-state floor read as "site drift" and
    blamed on the scraper. It is not drift; it is UNREACHABLE, exactly the LA
    distinction. So a workbook only counts as READ when the export answered 200
    with an actual xlsx; if NONE could be read, we record SOURCE_UNREACHABLE["CO"]
    and let the caller say so out loud instead of manufacturing a drift alarm on a
    scraper that is fine (it returns the full year from any un-blocked IP).
    """
    from openpyxl import load_workbook
    do_get = get or (lambda u, **kw: requests.get(u, **kw))
    landing = "https://cdle.colorado.gov/employers/layoff-separations/layoff-warn-list"
    ids = []
    try:
        resp = do_get(landing, headers=UA, timeout=TIMEOUT)
        if resp.status_code == 200:
            ids = list(dict.fromkeys(re.findall(r"docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]{20,})", resp.text)))
        else:
            print(f"CO landing returned HTTP {resp.status_code} (bot wall) — using pinned workbook IDs")
    except Exception as e:
        print(f"CO landing fetch failed: {e} — using pinned workbook IDs")
    # known workbooks (2026 current + 2025 archive) as fallback
    for known in ("19jmo4Cwj933cmSBKV1t0zZ5O-2H5IpiLIhSH9MF8WF0",
                  "1aFv4ntRhjnTMFKqBnuzbIkExCWgp6vnblYGm_h9GUeI"):
        if known not in ids:
            ids.append(known)
    out = []
    workbooks_read = 0        # exports that answered 200 with a real xlsx
    workbooks_blocked = 0     # transport error, non-200, or a non-xlsx body
    for sid in ids[:6]:
        try:
            xls = do_get(
                f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx",
                headers=UA, timeout=60)
            # A 403/429 rate-limit or an interstitial comes back as an HTML body,
            # not a zip. Only a 200 whose payload actually starts with the xlsx/zip
            # magic is a workbook we truly read; anything else is a refusal, and a
            # refusal must be counted as UNREACHABLE, never parsed as zero rows.
            if xls.status_code != 200 or xls.content[:2] != b"PK":
                workbooks_blocked += 1
                print(f"CO workbook {sid[:8]} blocked: HTTP {xls.status_code}, "
                      f"{len(xls.content)} bytes (not xlsx) — UNREACHABLE, not empty")
                continue
            wb = load_workbook(io.BytesIO(xls.content), read_only=True, data_only=True)
        except Exception as e:
            workbooks_blocked += 1
            print(f"CO workbook {sid[:8]} failed: {e} — UNREACHABLE, not empty")
            continue
        workbooks_read += 1
        for ws in wb.worksheets:
            if "warn" not in ws.title.lower() or "archiv" in ws.title.lower():
                continue
            rows = ws.iter_rows(values_only=True)
            headers = None
            for row in rows:
                vals = ["" if c is None else str(c).strip() for c in row]
                if headers is None:
                    if any("company" in v.lower() for v in vals):
                        headers = [v.lower() for v in vals]
                    continue
                rl = dict(zip(headers, vals))
                company = rl.get("company") or ""
                jobs = _count(rl.get("total notified") or rl.get("co notifications") or "")
                date = (_to_iso_date(rl.get("begin date") or "")
                        or _to_iso_date(rl.get("warn date") or "")
                        or _to_iso_date(rl.get("received") or ""))
                e = _entry("CO", company, jobs, date)
                if e:
                    out.append(e)
    # Distinguish UNREACHABLE from empty: if not one workbook could be read (every
    # export refused this runner) then a 0 here is a block, not a vanished state,
    # and the floor check must not blame the scraper for drift it did not commit.
    if workbooks_read == 0 and workbooks_blocked:
        SOURCE_UNREACHABLE["CO"] = (
            f"Google Sheets export refused this runner for all "
            f"{workbooks_blocked} workbook(s) (403/429 or non-xlsx body); "
            f"CO is UNREAD this run, not empty — the scraper reads the full "
            f"year from any un-blocked IP")
        print(f"::warning:: CO is UNREACHABLE this run: {SOURCE_UNREACHABLE['CO']}")
    return out


def _pdf_tables(content):
    import pdfplumber
    tables = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for t in page.extract_tables() or []:
                tables.append(t)
    return tables


def _pdf_text(content):
    import pdfplumber
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


def fetch_id():
    """Idaho: one cumulative text-layer PDF; link rotates, scraped from landing."""
    landing = "https://www.labor.idaho.gov/businesses/layoff-assistance/"
    page = requests.get(landing, headers=UA, timeout=TIMEOUT).text
    # The filename rotates. It carried a version suffix until mid-2026
    # (Idaho-WARN-Notices-3.2.pdf) and is now plain Idaho-WARN-Notices.pdf, in
    # an uploads folder dated the month they last republished. Match the stem,
    # then fall back to any WARN-named pdf on the page.
    m = re.search(r'href="(https://www\.labor\.idaho\.gov/wp-content/uploads/'
                  r'\d{4}/\d{2}/Idaho-WARN-Notices[^"]*\.pdf)"', page)
    if not m:
        m = re.search(r'href="(https?://[^"]*/wp-content/uploads/[^"]*'
                      r'[Ww][Aa][Rr][Nn][^"]*\.pdf)"', page)
    if not m:
        raise RuntimeError("ID: WARN pdf link not found on landing page")
    pdf = requests.get(m.group(1), headers=UA, timeout=60)
    out = []
    for table in _pdf_tables(pdf.content):
        for row in table:
            cells = ["" if c is None else str(c).strip() for c in row]
            if len(cells) < 9 or "company" in cells[2].lower():
                continue
            # Date of Letter | Updates | Company | Address | City | State | Zip | Affected | Effective
            jobs = _count(cells[7])
            date = _to_iso_date(cells[8]) or _to_iso_date(cells[0])
            if jobs <= 0:
                jobs = llm_count_from_text(" ".join(cells), f"ID {cells[2]}")
            e = _entry("ID", cells[2], jobs, date, cells[4], detail_url=m.group(1))
            if e:
                out.append(e)
    return out


def _la_entries_from_tables(tables, url):
    """LA row shapes: 2026+ is Company|Address|Notice|Layoff|Affected|Industry;
    2015-2025 is Company(+addr)|Notice|Layoff|Affected|Industry."""
    out = []
    for table in tables:
        for row in table:
            cells = ["" if c is None else str(c).strip().replace("\n", " ") for c in row]
            if len(cells) < 5 or "company" in (cells[0] or "").lower():
                continue
            if len(cells) >= 6:
                company, notice, layoff, affected = cells[0], cells[2], cells[3], cells[4]
            else:
                company, notice, layoff, affected = cells[0], cells[1], cells[2], cells[3]
            jobs = _count(affected)
            date = _to_iso_date(layoff) or _to_iso_date(notice)
            if jobs <= 0:
                jobs = llm_count_from_text(" ".join(cells), f"LA {company}")
            e = _entry("LA", company, jobs, date, detail_url=url)
            if e:
                out.append(e)
    return out


def fetch_la(get=None):
    """Louisiana: per-year text-layer PDFs, updated in place. laworks.net now
    hosts only the current two years — WarnNotices2015-2024.pdf all 404 on the
    live site (verified 2026-07-18) — so removed years fall back to a Wayback
    Machine snapshot taken after the year closed (same trust as MN's CDX
    discovery). Every year is fail-isolated.

    THE COUNT THIS STATE REPORTS IS ONLY MEANINGFUL IF THE ARCHIVE ANSWERED.
    Two of eleven years are live; the other nine ride entirely on
    web.archive.org. When that host does not answer, this function used to
    return the two live years and say nothing, and 33 notices against a floor
    of 324 read on the health page as "collapsed vs its own history — likely
    site drift", which is an accusation against a parser that is working.
    Measured on 2026-08-18: this returns 324 from a laptop and returned 33 from
    the runner the same morning, and 33 is exactly 21 (2025) + 12 (2026).

    So a year that is MISSING is distinguished from a year that was
    UNREACHABLE. A coherent 404 from Wayback means no snapshot of that year
    exists — a real gap in the archive, ours to solve by finding another route.
    A transport error or a 5xx/429 from web.archive.org means we never got to
    ask, and that is recorded in SOURCE_UNREACHABLE for the caller to say out
    loud instead of blaming the parser.

    `get` is injectable so the tests can drive both branches without a network.
    """
    from datetime import date as _date
    do_get = get or (lambda u, **kw: requests.get(u, **kw))
    out = []
    unreachable_years = []
    for y in range(2015, _date.today().year + 1):
        live = f"https://www.laworks.net/Downloads/WFD/WarnNotices{y}.pdf"
        content, source = None, live
        try:
            resp = do_get(live, headers=UA, timeout=60)
            if resp.status_code == 200 and resp.content[:4] == b"%PDF":
                content = resp.content
        except Exception:
            pass
        archive_answered = True
        if content is None:
            # April-after-year-end targets a complete-year snapshot while
            # steering clear of later capture dates that archived the 404
            wayback = f"https://web.archive.org/web/{y + 1}0401id_/{live}"
            archive_answered = False
            for attempt in range(2):  # Wayback is slow and flaky — retry once
                try:
                    resp = do_get(wayback, headers=UA, timeout=90)
                except Exception:
                    continue
                # A 429 or a 5xx is the archive refusing to serve this runner,
                # which is the same fact as a timeout wearing a status code.
                if resp.status_code == 429 or resp.status_code >= 500:
                    break
                archive_answered = True
                if resp.status_code == 200 and resp.content[:4] == b"%PDF":
                    content, source = resp.content, wayback
                break
        if content is None:
            if archive_answered:
                print(f"    LA {y}: no live PDF and no archived snapshot")
            else:
                unreachable_years.append(y)
                print(f"    LA {y}: web.archive.org did not answer — this "
                      f"year is UNREAD, not absent")
            continue
        try:
            out += _la_entries_from_tables(_pdf_tables(content), source)
        except Exception as exc:
            print(f"    LA {y}: PDF parse failed ({exc})")
    if unreachable_years:
        SOURCE_UNREACHABLE["LA"] = (
            f"web.archive.org did not answer for {len(unreachable_years)} "
            f"year(s) ({unreachable_years[0]}-{unreachable_years[-1]}); "
            f"only the {len(range(2015, _date.today().year + 1)) - len(unreachable_years)} "
            f"reachable year(s) are in this count")
        print(f"::warning:: LA is INCOMPLETE this run: "
              f"{SOURCE_UNREACHABLE['LA']}")
    return out


_NC_WARN_BASE = ("https://www.commerce.nc.gov/data-tools-reports/"
                 "labor-market-data-tools/workforce-warn-reports")


def _nc_year_csv(y):
    """Current NC vintage: Drupal year page -> rotating date-stamped CSV on
    files.nc.gov (the page slug only exists for the current year or two)."""
    import csv as _csv
    page = requests.get(f"{_NC_WARN_BASE}/report-workforce-warn-summary-list-{y}",
                        headers=UA, timeout=TIMEOUT)
    if page.status_code != 200:
        return []
    m = re.search(r'href="(https://files\.nc\.gov/[^"]+\.csv[^"]*)"', page.text)
    if not m:
        return []
    text = requests.get(_html.unescape(m.group(1)), headers=UA, timeout=TIMEOUT).text
    out = []
    for r in _csv.DictReader(io.StringIO(text)):
        jobs = _count(r.get("Number affected at this location") or "")
        date = (_to_iso_date(r.get("Effective Date") or "")
                or _to_iso_date(r.get("Date of Notice") or ""))
        e = _entry("NC", r.get("WARN Notice: WARN Notice Name"), jobs, date,
                   r.get("City") or "", kind=r.get("WARN notice type") or "",
                   detail_url=page.url)
        if e:
            out.append(e)
    return out


def _nc_header_cols(cells):
    """Map the heterogeneous NC header spellings (2018-2021 grids, 2022+
    grids/CSV) to field indexes; None unless the row looks like a header."""
    idx = {}
    for i, c in enumerate(cells):
        lc = re.sub(r"\s+", " ", str(c or "")).strip().lower()
        if "warn notice name" in lc or lc == "company":
            idx["company"] = i
        elif "number affected" in lc or "no. of employees" in lc:
            idx["jobs"] = i
        elif lc == "effective date":
            idx["eff"] = i
        elif lc in ("date of notice", "notice date"):
            idx["notice"] = i
        elif lc == "city":
            idx["city"] = i
        elif "warn notice type" in lc or lc.replace(" ", "") == "layoff/closure":
            idx["kind"] = i
    return idx if "company" in idx and "jobs" in idx else None


def _nc_plausible_company(name):
    """Reject a misaligned NC row before it can become an entry.

    The grid parser's header column-map persists across pages, and the 2015-2017
    x-position bucketing can drop a summary/continuation line into a data row --
    either way the company cell ends up holding a bare number ('0', '18'). Those
    rows were previously dropped only INCIDENTALLY (their count parses to 0, so
    _entry rejected them). Validating the company makes the skip deliberate,
    which is the precondition for running a count-fallback on NC: a rescued count
    must never be attached to a garbage row."""
    s = re.sub(r"\s+", " ", str(name or "")).strip()
    if not s:
        return False
    if re.fullmatch(r"[\d,.\s/-]+", s):     # bare number or date fragment
        return False
    # Two consecutive letters is the discriminator: it keeps every real name
    # (verified against all 1151 current NC rows, incl. "3M Company", "24 Hour
    # Fitness", "BP") while rejecting "0" / "18" / "1a" style misalignment.
    if not re.search(r"[A-Za-z]{2,}", s):
        return False
    return True


def _nc_grid_entries(tables, url):
    """Header-keyed rows from the gridded NC archive PDFs (2018+). The 2022+
    files only carry the header on page 1, so the column map persists across
    tables/pages."""
    out, cols = [], None
    for table in tables:
        for row in table:
            cells = ["" if c is None else re.sub(r"\s+", " ", str(c)).strip() for c in row]
            hdr = _nc_header_cols(cells)
            if hdr:
                cols = hdr
                continue
            if not cols:
                continue
            get = lambda k: cells[cols[k]] if k in cols and cols[k] < len(cells) else ""
            if not _nc_plausible_company(get("company")):
                continue  # stale column map hit a summary/continuation row
            jobs = _count(get("jobs"))
            if jobs <= 0:  # column shifted; the row is validated, so rescue is safe
                jobs = llm_count_from_text(" ".join(str(c) for c in cells),
                                           f"NC {get('company')}")
            e = _entry("NC", get("company"), jobs,
                       _to_iso_date(get("eff")) or _to_iso_date(get("notice")),
                       get("city"), kind=get("kind"), detail_url=url)
            if e:
                out.append(e)
    return out


# 2015-2017 NC reports have no table grid and interleave monthly summary
# blocks between data rows; these markers identify non-data lines.
_NC_TEXT_NOISE = re.compile(
    r"WARN Notice|Date Range|Effective Date|Month:|Total Notices|Sum of|"
    r"Layoff:|Closure:|Grand Total")


def _nc_text_rows(words, url):
    """2015-2017 NC vintage: bucket pdfplumber words into columns using the
    x-positions of the per-page header row (Notice Date | Effective Date |
    Company Name | City | # Emp. Affected + kind). Wrapped company/city
    continuation lines merge into the previous row."""
    anchors = {}
    for w in words:
        if w["text"] in ("Notice", "Effective", "Company", "City", "#"):
            anchors.setdefault(round(w["top"]), {}).setdefault(w["text"], w["x0"])
    header_top = cols = None
    for top, d in sorted(anchors.items()):
        if {"Notice", "Effective", "Company", "City"} <= set(d):
            header_top = top
            cols = [d["Notice"], d["Effective"], d["Company"], d["City"],
                    d.get("#", d["City"] + 55)]
            break
    if cols is None:
        return []
    lines, cur_top = [], None
    for w in sorted(words, key=lambda x: (x["top"], x["x0"])):
        if w["top"] <= header_top + 2:
            continue
        if cur_top is None or w["top"] - cur_top > 3:
            lines.append([])
            cur_top = w["top"]
        lines[-1].append(w)
    rows = []  # [notice, effective, company, city, count+kind tail]
    for ws in lines:
        cells = ["", "", "", "", ""]
        for w in sorted(ws, key=lambda x: x["x0"]):
            ci = 0
            for i, cx in enumerate(cols):
                if w["x0"] >= cx - 6:
                    ci = i
            cells[ci] = f"{cells[ci]} {w['text']}".strip()
        if _NC_TEXT_NOISE.search(" ".join(cells)):
            continue
        if _to_iso_date(cells[0]) or _to_iso_date(cells[1]):
            rows.append(cells)
        elif rows:  # wrapped continuation of the previous row
            for i in (2, 3):
                if cells[i]:
                    rows[-1][i] = f"{rows[-1][i]} {cells[i]}".strip()
    out = []
    for cells in rows:
        if not _nc_plausible_company(cells[2]):
            continue  # x-position bucketing mislanded a summary/continuation line
        kind_m = re.search(r"Layoffs?|Closures?", cells[4])
        jobs = _count(cells[4])
        if jobs <= 0:  # row is validated, so a count rescue can't hit garbage
            jobs = llm_count_from_text(" ".join(cells), f"NC {cells[2]}")
        e = _entry("NC", cells[2], jobs,
                   _to_iso_date(cells[1]) or _to_iso_date(cells[0]), cells[3],
                   kind=kind_m.group(0) if kind_m else "", detail_url=url)
        if e:
            out.append(e)
    return out


def fetch_nc():
    """NC Commerce, three vintages (verified 2026-07-18): the current year or
    two live on Drupal pages with a rotating CSV; 2015-2025 hang off the
    warn-summary-report-archives page as per-year PDFs — gridded tables for
    2018+, headerless text columns for 2015-2017. Everything fail-isolated."""
    from datetime import date as _date
    out = []
    for y in (_date.today().year, _date.today().year - 1):  # Jan rollover safety
        try:
            out += _nc_year_csv(y)
        except Exception as exc:
            print(f"    NC {y}: summary-list failed ({exc})")
    try:
        arch = requests.get(f"{_NC_WARN_BASE}/warn-summary-report-archives",
                            headers=UA, timeout=TIMEOUT)
        hrefs = re.findall(r'href="([^"]*warn[^"]*/open)"', arch.text, re.I)
    except Exception as exc:
        print(f"    NC: archive index failed ({exc})")
        hrefs = []
    seen = set()
    for href in hrefs:
        ym = re.search(r"(20\d\d)", href)
        # 2014 exists on the page but predates the tracker's 2015 cutoff
        if not ym or int(ym.group(1)) < 2015 or href in seen:
            continue
        seen.add(href)
        url = href if href.startswith("http") else "https://www.commerce.nc.gov" + href
        try:
            resp = requests.get(url, headers=UA, timeout=60)
            if resp.status_code != 200 or resp.content[:4] != b"%PDF":
                raise RuntimeError(f"HTTP {resp.status_code} / not a PDF")
            entries = _nc_grid_entries(_pdf_tables(resp.content), url)
            if not entries:  # 2015-2017 vintage
                import pdfplumber
                with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                    for p in pdf.pages:
                        entries += _nc_text_rows(p.extract_words(), url)
            out += entries
        except Exception as exc:
            print(f"    NC {ym.group(1)} archive: failed ({exc})")
    return out


# Nevada's master PDF glues tokens ("209SK Food Group"), so rows are parsed
# from text lines; the trailing County token is stripped via this vocabulary.
_NV_COUNTIES = (
    "Carson City|Churchill|Clark|Douglas|Elko|Esmeralda|Eureka|Humboldt|Lander|"
    "Lincoln|Lyon|Mineral|Nye|Pershing|Storey|Washoe|White Pine|Statewide|Remote|Various"
)
_NV_ROW = re.compile(
    r"^(?:(?P<recv>\d{1,2}/\d{1,2}/\d{2,4})|Unknown|Multiple)?\s*"
    r"(?:(?P<eff>\d{1,2}/\d{1,2}/\d{2,4})|Unknown|Multiple)?\s*"
    r"(?P<kind>Layoffs?|Closures?)\s*"
    r"(?P<jobs>[\d,]+|Unknown|NR)\s*"
    r"(?P<rest>.+?)\s*"
    r"(?P<warn>Non-?WARN|WARN)?$")

# Known NV municipalities, longest first — the text line is
# "<company> <city> <county>" with no delimiters, so the split must be
# vocabulary-driven ("Spirit Airlines Las Vegas Clark" is genuinely ambiguous
# to a regex: company names may end in place-like words).
_NV_CITIES = [
    "North Las Vegas", "Incline Village", "Boulder City", "Carson City", "Sun Valley",
    "Battle Mountain", "West Wendover", "Spring Creek", "Indian Springs", "Mound House",
    "Las Vegas", "Henderson", "Reno", "Sparks", "Stateline", "Primm", "Elko",
    "Mesquite", "Fernley", "Minden", "Gardnerville", "Winnemucca", "Fallon",
    "Pahrump", "Laughlin", "Verdi", "Sloan", "Jean", "McCarran", "Ely", "Yerington",
    "Lovelock", "Tonopah", "Hawthorne", "Caliente", "Wells", "Jackpot", "Dayton",
    "Silver Springs", "Moapa", "Overton", "Amargosa Valley", "Crystal Bay", "Genoa",
    "Zephyr Cove", "Round Mountain", "Eureka", "Remote", "Statewide", "Various",
]

# Alternations are built longest-first so "North Las Vegas" always wins over
# "Las Vegas" (Python alternation is leftmost-first, so ordering IS the
# precedence rule and sorting by length makes it independent of hand-ordering).
_NV_CITY_ALT = "|".join(re.escape(c) for c in sorted(_NV_CITIES, key=len, reverse=True))

# Both strips tolerate a MISSING space before the place text. pdfplumber glues
# the city and county together on multi-site notices, which is how
# "Spirit Airlines Las Vegas/Reno Clark/Washoe" reached the live tracker stored
# as the employer "Spirit Airlines Las Vegas/RenoClark/Washoe": the county strip
# wanted \s+ before "Clark", found "...Reno" instead, so nothing was stripped
# and the whole line became the company name. A "/"-joined run of cities or
# counties is one place, not several.
_NV_COUNTY_RX = re.compile(
    r"(?:\s+|(?<=[a-z]))(?:%s)(?:\s*/\s*(?:%s))*\s*$" % (_NV_COUNTIES, _NV_COUNTIES))
_NV_CITY_RX = re.compile(
    r"(?:\s+|^)((?:%s)(?:\s*/\s*(?:%s))*)\s*$" % (_NV_CITY_ALT, _NV_CITY_ALT))


def _nv_place_split(rest):
    """Split the NV master PDF's "<company> <city> <county>" run into
    (company, city). The PDF carries no delimiters, so the split is
    vocabulary-driven.

    A city the vocabulary does not know leaves the company UNCHANGED and the
    city blank, and `fetch_nv` reports the count. That is deliberate: guessing
    would mean stripping any trailing place-like word, and truncating a real
    employer ("... Enterprise", "... Paradise") is a worse, less visible
    failure than an empty city column. The caller's diagnostic is what turns
    the vocabulary's ceiling from silent into observable.
    """
    rest = _NV_COUNTY_RX.sub("", (rest or "").strip()).strip()
    m = _NV_CITY_RX.search(rest)
    if not m:
        return rest, ""
    city = re.sub(r"\s*/\s*", "/", m.group(1).strip())
    return rest[: m.start()].strip(), city


# Nevada DETR now sits behind an Akamai bot-wall: the /Page/WARN landing HTML
# 403s for non-browsers, so link discovery is dead. The cumulative master PDF
# under /content/media/ IS reachable, but ONLY with a full browser-like header
# set (User-Agent alone still 403s). Current year lives at a stable base
# filename; prior years are date-suffixed archives. A 0-notice NV result now
# means "check the master filename", not a silent failure (the health tripwire
# already flags it).
_NV_HEADERS = {
    "User-Agent": UA["User-Agent"],
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Upgrade-Insecure-Requests": "1",
}
# Bluehost mirrors DETR's master PDF into uploads on a daily WP-cron (its IP is
# not Akamai-blocked, unlike CI's data-center IP). Read the mirror FIRST — it is
# reachable from CI — and fall back to DETR direct for residential runs or if the
# mirror is ever missing. The first URL that returns a valid PDF wins.
_NV_MASTER_PDFS = (
    "https://asktherecruiter.com/blog/wp-content/uploads/nv-warn-master.pdf",
    "https://detr.nv.gov/content/media/WARN_and_Non_WARN_Master_w_Logo.pdf",
)

def fetch_nv():
    """Nevada DETR cumulative master PDF. Landing page is Akamai-403 (no link
    discovery); fetch the known master URL directly with browser headers."""
    import pdfplumber
    out = []
    for url in _NV_MASTER_PDFS:
        try:
            resp = requests.get(url, headers=_NV_HEADERS, timeout=TIMEOUT)
        except requests.RequestException:
            continue
        if resp.status_code != 200 or not resp.content.startswith(b"%PDF"):
            continue
        before = len(out)
        unplaced = []
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                for line in (page.extract_text() or "").splitlines():
                    line = line.strip()
                    if "Employer" in line or not re.search(r"Layoffs?|Closures?", line):
                        continue
                    m = _NV_ROW.match(line)
                    if not m:
                        continue
                    company, city = _nv_place_split(m.group("rest"))
                    e = _entry("NV", company, _count(m.group("jobs")),
                               _to_iso_date(m.group("eff") or "") or _to_iso_date(m.group("recv") or ""),
                               city, kind=m.group("kind"), detail_url=url)
                    if e:
                        out.append(e)
                        if not city:
                            unplaced.append(e["company_name"])
        # The master PDF is cumulative, so the first source that yields rows has
        # everything; stop so a residential run does not parse the mirror AND
        # DETR and double the list.
        if len(out) > before:
            # Every NV line carries a city, so an unresolved one means the place
            # text is still glued to the employer name. Say so: a thin
            # vocabulary that fails silently is how a wrong company name reaches
            # the public tracker and nobody finds out.
            if unplaced:
                print(f"NV: {len(unplaced)} of {len(out) - before} notices had no "
                      f"recognisable city, so place text may remain in the employer "
                      f"name: {'; '.join(unplaced[:8])}")
            break
    return out


# Radware bot-walls mn.gov HTML pages but NOT /deed/assets/*.pdf — discovery
# of the unpredictable tcm1045 ids falls back to the Wayback CDX index, and a
# seed list keeps known history importable through a total discovery outage.
_MN_SEED_PDFS = [  # verified-live 2026-07-19; CDX discovery adds the rest
    "https://mn.gov/deed/assets/plant-closing-mass-layoff-warn-report-2023_tcm1045-663809.pdf",
    "https://mn.gov/deed/assets/plant-closing-mass-layoff-warn-october-2025_tcm1045-712065.pdf",
    "https://mn.gov/deed/assets/plant-closing-mass-layoff-warn-2026-january_tcm1045-722872.pdf",
    # 2026 monthly reports the CDX index still lacks (verified-live 2026-08-24):
    # DEED published them but Wayback has not archived their asset URLs, and the
    # HTML index that lists them is Radware/ShieldSquare CAPTCHA-walled (we do
    # not bypass it), so seeding is the only reach. Both parse with the existing
    # _mn_parse_table (+8 notices dry-run 2026-08-24). Discovered via a search
    # index of mn.gov/deed/assets, which DOES index these PDFs.
    "https://mn.gov/deed/assets/plant-closing-mass-layoff-warn-2026-february_tcm1045-742362.pdf",
    "https://mn.gov/deed/assets/plant-closing-mass-layoff-warn-2026-april_tcm1045-749441.pdf",
    "https://mn.gov/deed/assets/plant-closing-mass-layoff-warn-2026-june_tcm1045-758364.pdf",
    # keep recent months here until CDX archives them or a JS-render discovery lands
]


def _mn_parse_table(table):
    """(company, jobs, date, city) rows from one MN report table, template-agnostic.

    DEED changed the monthly template mid-2026: multi-word headers now wrap
    across TWO physical rows ('Affected' / 'Workers'), a leading blank column
    and a trailing 'Layoff Count' column appear, and 'Account: City' became
    'City'. The old parser matched whole header strings in single cells, so it
    silently returned 0 rows on every new-template report. This merges a wrapped
    continuation header row before keyword-matching, handling both layouts."""
    def cell(row, j):
        return re.sub(r"\s+", " ", (row[j] or "")).strip() if j < len(row) else ""
    hdr_i = next((i for i, r in enumerate(table[:3])
                  if any("Layoff Name" in (c or "") for c in r)), None)
    if hdr_i is None:
        return []
    ncol = len(table[hdr_i])
    combined = []
    for j in range(ncol):
        h = cell(table[hdr_i], j)
        if hdr_i + 1 < len(table):
            nxt = cell(table[hdr_i + 1], j)
            if nxt and not re.search(r"\d", nxt) and len(nxt) <= 12 and not _to_iso_date(nxt):
                h = f"{h} {nxt}".strip()
        combined.append(h.lower())

    def find(*keys):
        return next((j for j, h in enumerate(combined) if any(k in h for k in keys)), None)
    ci_name = find("layoff name")
    ci_city = find("city")
    ci_date = find("layoff start")
    ci_recv = find("warn received")
    ci_jobs = find("affected worker")
    if ci_name is None or ci_jobs is None:
        return []
    start = hdr_i + 1
    if start < len(table):
        first = [cell(table[start], j) for j in range(ncol)]
        if not any(re.search(r"\d", x) for x in first):
            start += 1  # skip the wrapped-continuation header row
    out = []
    for row in table[start:]:
        cells = [cell(row, j) for j in range(ncol)]
        name = re.sub(r"\s+20\d\d$", "", cells[ci_name] if ci_name < len(cells) else "")
        if not name or name.lower().startswith(("total", "rr start", "count")):
            continue
        jobs = _count(cells[ci_jobs]) if ci_jobs < len(cells) else 0
        date = _to_iso_date(cells[ci_date]) if (ci_date is not None and ci_date < len(cells)) else ""
        if not date and ci_recv is not None and ci_recv < len(cells):
            date = _to_iso_date(cells[ci_recv])
        city = cells[ci_city] if (ci_city is not None and ci_city < len(cells)) else ""
        out.append((name, jobs, date, city))
    return out


def fetch_mn():
    """MN DEED monthly/annual report PDFs (assets bypass the bot wall).
    Discovery = seed list + Wayback CDX; parsing via template-agnostic table map."""
    import pdfplumber
    urls = set(_MN_SEED_PDFS)
    for attempt in range(2):  # Wayback CDX is slow and flaky — retry once
        try:
            cdx = requests.get(
                "https://web.archive.org/cdx/search/cdx",
                params={"url": "mn.gov/deed/assets/plant-closing*",
                        "collapse": "urlkey", "fl": "original", "limit": "500"},
                headers=UA, timeout=90)
            for u in cdx.text.split():
                if re.search(r"/deed/assets/plant-closing[^\s]*_tcm1045-\d+\.pdf$", u):
                    urls.add(u.replace("http://", "https://"))
            break
        except Exception as exc:  # discovery outage -> seeds still import
            print(f"    MN: CDX discovery failed ({exc}); "
                  + ("retrying" if attempt == 0 else "using seed list"))
    out = []
    for url in sorted(urls):
        try:
            resp = requests.get(url, headers=UA, timeout=TIMEOUT)
        except Exception:
            continue
        if resp.status_code != 200 or b"%PDF" not in resp.content[:1024]:
            continue
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    for company, jobs, date, city in _mn_parse_table(table):
                        e = _entry("MN", company, jobs, date, city, detail_url=url)
                        if e:
                            out.append(e)
    return out


# mass.gov (Akamai) rejects HTTP/1.1 outright and demands a full Chrome header
# fingerprint — hence httpx with http2 rather than requests (recon-verified:
# same headers 403 over HTTP/1.1, 200 over HTTP/2).
_MA_HEADERS = {
    "User-Agent": UA["User-Agent"],
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none", "Sec-Fetch-User": "?1",
    "sec-ch-ua": '"Chromium";v="126", "Not/A)Brand";v="8"',
    "sec-ch-ua-mobile": "?0", "sec-ch-ua-platform": '"macOS"',
    "Upgrade-Insecure-Requests": "1",
}
_MA_LANDING = ("https://www.mass.gov/info-details/worker-adjustment-and-retraining-"
               "notification-act-warn-layoff-and-closure-updates")

def _ma_first_date(value):
    """First parseable date from free text ('8/15/26 & 11/30/26', ranges, serials)."""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    for token in re.split(r"\s*(?:-|&|,|\band\b|\bending\b|\bthrough\b)\s*", str(value or "")):
        d = _to_iso_date(token)
        if d:
            return d
    return None

def fetch_ma():
    """Massachusetts: weekly CSV (current FY) + one XLSX per archived FY."""
    import httpx
    from openpyxl import load_workbook
    client = httpx.Client(http2=True, headers=_MA_HEADERS, follow_redirects=True, timeout=TIMEOUT)
    landing = client.get(_MA_LANDING)
    out = []

    def add(company, jobs_raw, date_raw, fallback_date, city):
        company = re.sub(r"^\*Updated\*\s*", "", str(company or ""))
        city = re.sub(r",\s*MA$", "", str(city or "")).strip()
        date = _ma_first_date(date_raw) or _ma_first_date(fallback_date)
        e = _entry("MA", company, _count(str(jobs_raw or "")), date, city, detail_url=_MA_LANDING)
        if e:
            out.append(e)

    csv_urls = re.findall(r'href="(https://www\.mass\.gov/files/csv/[^"]+\.csv)"', landing.text)
    import csv as _csv
    for url in csv_urls[:1]:
        r = client.get(_html.unescape(url))
        if r.status_code == 200:
            try:  # the CSV mixes encodings ("Labouré College" is cp1252)
                text = r.content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = r.content.decode("cp1252", errors="replace")
            for row in _csv.DictReader(io.StringIO(text)):
                add(row.get("EMPLOYER"), row.get("# EMPLOYEES IMPACTED"),
                    row.get("DATE(S) OF LAYOFFS"), row.get("RECEIVED"), row.get("CITY/TOWN"))

    fy_urls = re.findall(r'href="(https://www\.mass\.gov/doc/fy\d\d-warn-report[^"]*/download)"', landing.text)
    for url in sorted(set(fy_urls)):
        r = client.get(_html.unescape(url))
        if r.status_code != 200:
            continue
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        for ws in wb.worksheets:
            header, cols = None, {}
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else c for c in row]
                if header is None:
                    joined = " ".join(str(c).upper() for c in cells)
                    if "EMPLOYER" in joined or "COMPANY NAME" in joined:
                        header = [str(c).strip().upper() for c in cells]
                        for i, h in enumerate(header):
                            if "EMPLOYER" in h or "COMPANY" in h: cols["company"] = i
                            elif "IMPACT" in h or "AFFECT" in h:  cols["jobs"] = i
                            elif "LAYOFF" in h and "DATE" in h.replace("(S)", "S"): cols["date"] = i
                            elif h in ("RECEIVED", "DATE RECEIVED"): cols["recv"] = i
                            elif "CITY" in h:                     cols["city"] = i
                    continue
                if "company" not in cols or "jobs" not in cols:
                    continue
                get = lambda k: cells[cols[k]] if k in cols and cols[k] < len(cells) else ""
                if not str(get("company")).strip():
                    continue
                add(get("company"), get("jobs"), get("date"), get("recv"), get("city"))
    client.close()
    return out


# --- New York history: the retired dol.ny.gov database (frozen 4/1/2025) ---
# The hub page itself lists the final Jan-Apr 2025 notices and links plain-
# HTML year archives for 2023/2024. All three share one 4-column table
# (Company | Region | Date Posted | Notice Dated) with NO headcount — the
# count and layoff date live in the per-notice PDF behind each company link
# (verified 2026-07-18). Live NY data comes from warn-scraper's dashboard
# scraper; identical dedup hashes make the overlap upsert-safe.
_NY_HUB = "https://dol.ny.gov/warn-notices"


def _ny_listing_rows(page_html):
    """(company, detail_url, notice_date_text) rows from a retired-NY listing."""
    rows = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", page_html, re.S | re.I):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)
        if len(cells) < 4:
            continue
        link = re.search(r'href="\s*(/warn-[^"]+)"', cells[0])
        company = _strip_tags(cells[0])
        if not link or not company:
            continue
        rows.append((company, "https://dol.ny.gov" + link.group(1).strip(),
                     _strip_tags(cells[3])))
    return rows


def _ny_fields_from_text(text):
    """(jobs, layoff_date_iso) from the uniform NY WARN-unit notice layout.
    Dates wrap across lines in the PDF text layer ('February\\n12, 2024'), so
    the field windows are whitespace-flattened first. Amended notices list the
    original count before the '• Amended to N' bullet — the first-number parse
    keeps the original, matching the tracker's lower-bound convention."""
    flat = re.sub(r"\s+", " ", text or "")
    jm = (re.search(r"Total Number of Affected Workers\s*:?\s*([\d,]+)", flat, re.I)
          or re.search(r"Number of Affected Employees at Site\s*:?\s*([\d,]+)", flat, re.I)
          # Old 2022-2023 WARN-unit form used a single, differently-worded label.
          # "Number Affected" (Affected right after Number) can't collide with the
          # new form's "Total Number of Affected Workers" ("Number of Affected").
          or re.search(r"\bNumber Affected\s*:?\s*([\d,]+)", flat, re.I))
    jobs = _count(jm.group(1)) if jm else 0
    if jobs <= 0:  # unrecognized form; LLM tail-net reads the notice text
        jobs = llm_count_from_text(flat, "NY")
    dm = re.search(r"(?:Layoff|Closure)[^:]{0,15}Start Date\s*:\s*(.{0,220})", flat, re.I)
    date = _to_iso_date(dm.group(1)) if dm else ""
    if not date:
        nm = re.search(r"Date of Notice\s*:\s*(.{0,60})", flat, re.I)
        date = _to_iso_date(nm.group(1)) if nm else ""
    return jobs, date


def fetch_ny_history():
    """History backfill for NY from the retired database: walk each frozen
    listing, then fetch every linked per-notice PDF for its count/start date
    (falling back to the listing's notice date). Listings and notices are each
    fail-isolated."""
    listings = [("2025 hub", _NY_HUB),
                ("2024", "https://dol.ny.gov/2024-warn-notices"),
                ("2023", "https://dol.ny.gov/2023-warn-notices")]
    out = []
    for label, listing_url in listings:
        try:
            page = requests.get(listing_url, headers=UA, timeout=TIMEOUT)
            if page.status_code != 200:
                raise RuntimeError(f"HTTP {page.status_code}")
            rows = _ny_listing_rows(page.text)
        except Exception as exc:
            print(f"    NY {label}: listing failed ({exc})")
            continue
        kept = failed = 0
        for company, detail_url, notice_text in rows:
            try:
                resp = requests.get(detail_url, headers=UA, timeout=60)
                if resp.status_code != 200:
                    raise RuntimeError(f"HTTP {resp.status_code}")
                text = (_pdf_text(resp.content) if resp.content[:4] == b"%PDF"
                        else _strip_tags(resp.text))
                jobs, date = _ny_fields_from_text(text)
            except Exception:
                failed += 1
                continue
            e = _entry("NY", company, jobs, date or _to_iso_date(notice_text),
                       detail_url=detail_url)
            if e:
                out.append(e)
                kept += 1
        print(f"    NY {label}: {kept} kept, {failed} failed of {len(rows)} rows")
    return out


def fetch_ky():
    """Kentucky (kyworks.ky.gov): current + prior-year WARN xlsx workbooks linked
    from the Rapid Response page. Migrated off the generic warn-scraper tier,
    which only ships KY's frozen historical CSV (ends ~2021) and can't read KY's
    rotating, date-stamped xlsx filenames. Columns (verified 2026-07): company
    'Company: Company Name', jobs 'Number of Employees Affected', effective
    'Projected Date' (Excel serial), notice 'Date Received', county 'County',
    kind 'Closure or Layoff?', link 'Notice URL'."""
    from openpyxl import load_workbook
    from datetime import datetime as _dt, timedelta as _td
    landing = "https://kyworks.ky.gov/Services/Pages/Rapid-Response-Layoffs-and-Closures.aspx"
    origin = "https://kyworks.ky.gov"
    hrefs = []
    try:
        page = requests.get(landing, headers=UA, timeout=TIMEOUT).text
        hrefs = re.findall(r'href="(/Services/Documents/[^"]+\.xlsx)"', page, re.I)
    except Exception as e:
        print(f"KY landing fetch failed: {e}")
    if not hrefs:  # fallback to the known filename shapes
        hrefs = ["/Services/Documents/Website-WARN%20Notice%20Report%2007012026.xlsx",
                 "/Services/Documents/Prior%20Year%20Warn%20Notices.xlsx"]

    def _iso(v):
        if v is None:
            return ""
        if isinstance(v, _dt):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, (int, float)):  # Excel serial (epoch 1899-12-30)
            try:
                return (_dt(1899, 12, 30) + _td(days=int(v))).strftime("%Y-%m-%d")
            except Exception:
                return ""
        return _to_iso_date(str(v))

    out, seen = [], set()
    for href in dict.fromkeys(hrefs):
        url = href if href.lower().startswith("http") else origin + href
        try:
            xls = requests.get(url, headers=UA, timeout=60)
            wb = load_workbook(io.BytesIO(xls.content), read_only=True, data_only=True)
        except Exception as e:
            print(f"KY workbook fetch failed ({url[:60]}): {e}")
            continue
        for ws in wb.worksheets:
            headers = None
            for row in ws.iter_rows(values_only=True):
                cells = list(row)
                if headers is None:
                    lower = [str(c).strip().lower() if c is not None else "" for c in cells]
                    if any("company" in v for v in lower):
                        headers = lower
                    continue
                rl = dict(zip(headers, cells))

                def g(*names):
                    for n in names:
                        for k, v in rl.items():
                            if n in k:
                                return v
                    return ""
                company = str(g("company") or "").strip()
                jobs = _count(str(g("number of employees affected", "employees affected", "affected") or ""))
                date = _iso(g("projected date")) or _iso(g("date received", "received"))
                county = str(g("county") or "").strip()
                kind = str(g("closure or layoff", "layoff?") or "").strip()
                detail = str(g("notice url", "url") or "").strip()
                if not detail.lower().startswith("http"):
                    detail = ""
                key = (company.lower(), date, jobs, county.lower())
                if key in seen:
                    continue
                seen.add(key)
                e = _entry("KY", company, jobs, date, city=county, kind=kind, detail_url=detail)
                if e:
                    out.append(e)
    return out


#: States whose fetcher could not REACH one of its source documents this run,
#: as {STATE: why}. Populated during a sweep and cleared at the start of the
#: next one. This is the channel that keeps "we could not read it" from being
#: reported as "the site drifted" — see fetch_la. It is deliberately NOT an
#: exception: the reachable years are still real notices worth upserting, and
#: throwing them away to make the failure loud would trade data for a message.
SOURCE_UNREACHABLE = {}


CUSTOM_STATES = {
    "TX": fetch_tx, "FL": fetch_fl, "GA": fetch_ga, "OH": fetch_oh,
    "MI": fetch_mi, "CO": fetch_co, "ID": fetch_id, "LA": fetch_la,
    "NC": fetch_nc, "NV": fetch_nv, "MN": fetch_mn, "MA": fetch_ma,
    "NY": fetch_ny_history, "KY": fetch_ky,
}


def pull_warn_custom(states):
    """Fetch the custom states (intersection with `states`, or all when 'all')."""
    SOURCE_UNREACHABLE.clear()
    scrape_all = len(states) == 1 and str(states[0]).lower() == "all"
    wanted = list(CUSTOM_STATES) if scrape_all else [s.upper() for s in states if s.upper() in CUSTOM_STATES]
    results = []
    for st in wanted:
        try:
            entries = CUSTOM_STATES[st]()
            print(f"WARN {st} (custom): {len(entries)} notices kept")
            results.extend(entries)
        except Exception as e:
            print(f"WARN {st} (custom) FAILED: {e}")
    return results
